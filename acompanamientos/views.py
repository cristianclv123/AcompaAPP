import os
import openpyxl
import unicodedata
from django.shortcuts import render
from django.conf import settings

# --- FUNCIÓN AUXILIAR PARA NORMALIZAR Y LIMPIAR TEXTO ---
def normalizar_texto(texto):
    if not texto:
        return ""
    texto = str(texto).lower().strip()
    texto = texto.replace('.', '').replace(',', '').replace('-', '')
    texto_normalizado = "".join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    )
    return " ".join(texto_normalizado.split())

# --- FUNCIÓN DE COINCIDENCIA DE NOMBRES ---
def coinciden_nombres(query_normalizada, celda_normalizada):
    if not query_normalizada or not celda_normalizada:
        return False
    return query_normalizada in celda_normalizada or celda_normalizada in query_normalizada

# --- FUNCIONES DE DETECCIÓN PARA HOJAS AFL ---
def contiene_mes(texto):
    if not texto:
        return False
    meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio', 'julio',
             'agosto', 'septiembre', 'setiembre', 'octubre', 'noviembre', 'diciembre']
    texto_normalizado = normalizar_texto(texto)
    return any(mes in texto_normalizado for mes in meses)

def fila_es_ciclo(fila):
    if not fila:
        return False
    return any('ciclo' in normalizar_texto(celda) for celda in fila if celda)

def fila_es_grupo(fila):
    if not fila:
        return False
    return any('grupo' in normalizar_texto(celda) for celda in fila if celda)

def fila_es_dia(fila):
    if not fila:
        return False
    return any('dia' in normalizar_texto(celda) for celda in fila if celda)

def es_hoja_afl(nombre_hoja, filas_muestra, texto_cabecera, sheet_idx=None, rango_afl=None):
    if rango_afl is not None and sheet_idx is not None and sheet_idx in rango_afl:
        return True

    nombre_normalizado = normalizar_texto(nombre_hoja)
    encabezado_normalizado = normalizar_texto(texto_cabecera)

    if 'kawak' in nombre_normalizado or 'control' in nombre_normalizado:
        return False
    if 'cancha' in nombre_normalizado or 'distribucion' in encabezado_normalizado:
        return False

    if 'afl' in nombre_normalizado:
        return True
    if 'ciclo' in encabezado_normalizado or 'grupo' in encabezado_normalizado:
        return True
    if contiene_mes(nombre_hoja) and 'grupo' in encabezado_normalizado:
        return True

    for idx, fila in enumerate(filas_muestra):
        if fila_es_grupo(fila) and idx > 0:
            anterior = filas_muestra[idx - 1]
            if fila_es_ciclo(anterior) or fila_es_dia(anterior) or any(contiene_mes(celda) for celda in anterior if celda):
                return True
    return False

def calcular_rango_afl(sheetnames):
    nombres_norm = [normalizar_texto(nombre) for nombre in sheetnames]
    inicio = -1
    fin = -1

    for idx, nombre in enumerate(nombres_norm):
        if 'cancha' in nombre and ('2' in nombre or 'segundo' in nombre or '2do' in nombre or '2º' in nombre):
            inicio = idx
            break

    if inicio == -1:
        return None

    for idx in range(inicio + 1, len(nombres_norm)):
        nombre = nombres_norm[idx]
        if 'cancha' in nombre and 'afl' not in nombre:
            fin = idx
            break

    if fin > inicio + 1:
        return range(inicio + 1, fin)
    return None

def consulta(request):
    query_raw = request.GET.get('q', '').strip()
    query = normalizar_texto(query_raw)
    turnos_zonas = []
    turnos_afl = []
    
    ruta_excel = os.path.join(settings.BASE_DIR, 'Acompañamientos 5º-11º.xlsx')
    cancha_data = []

    colores = ['blue', 'green', 'orange']
    iconos = ['fa-volleyball', 'fa-mug-hot', 'fa-road', 'fa-location-dot']

    if os.path.exists(ruta_excel):
        try:
            wb = openpyxl.load_workbook(ruta_excel, data_only=True)
            rango_afl = calcular_rango_afl(wb.sheetnames)

            # ==============================================================
            # SECCIÓN QUE EXTRAE ÚNICAMENTE LAS DOS TABLAS SOLICITADAS
            # ==============================================================
            target_sheets = []
            
            for nm in wb.sheetnames:
                # AQUÍ ESTÁ EL CAMBIO: Solo permite exactamente la hoja que se llame "Cancha"
                if nm.strip().lower() == 'cancha':
                    if nm not in target_sheets:
                        target_sheets.append(nm)

            for ts in target_sheets:
                hoja_tmp = wb[ts]
                filas_tmp = list(hoja_tmp.iter_rows(values_only=True))

                filas_limpias = []
                for r in filas_tmp:
                    row_clean = []
                    for c in r:
                        val = "" if c is None else str(c).strip()
                        val_lower = val.lower()
                        if val_lower == 'none':
                            row_clean.append("")
                        else:
                            row_clean.append(val)
                    filas_limpias.append(row_clean)

                filas_con_datos = [r for r in filas_limpias if any(c != "" for c in r)]

                if filas_con_datos:
                    max_cols_tmp = max(len(r) for r in filas_con_datos)
                    cols_validas = []
                    for col_idx in range(max_cols_tmp):
                        if any(col_idx < len(r) and r[col_idx] != "" for r in filas_con_datos):
                            cols_validas.append(col_idx)

                    grupos_columnas = []
                    grupo_actual = []
                    for col in cols_validas:
                        if not grupo_actual:
                            grupo_actual.append(col)
                        else:
                            if col == grupo_actual[-1] + 1:
                                grupo_actual.append(col)
                            else:
                                grupos_columnas.append(grupo_actual)
                                grupo_actual = [col]
                    if grupo_actual:
                        grupos_columnas.append(grupo_actual)

                    for idx, grupo in enumerate(grupos_columnas):
                        filas_trim = []
                        for r in filas_con_datos:
                            r_padded = r + [""] * (max_cols_tmp - len(r))
                            row_final = [r_padded[i] for i in grupo]
                            if any(c != "" for c in row_final):
                                filas_trim.append(row_final)
                        
                        if filas_trim:
                            # AQUÍ ASIGNAMOS LOS TÍTULOS EXACTOS QUE PEDISTE A CADA TABLA
                            if idx == 0:
                                nombre_seccion = "DISTRIBUCIÓN DE LA CANCHA GRANDE DE 5º -8º SEGUNDO PERIODO"
                            elif idx == 1:
                                nombre_seccion = "DISTRIBUCIÓN DE CANCHA PARA LOS DÍAS MARTES"
                            else:
                                nombre_seccion = f"Tabla extra {idx + 1}"

                            cancha_data.append({
                                'sheet': nombre_seccion, 
                                'rows': filas_trim, 
                                'cols': len(grupo)
                            })

            # ==============================================================
            # BÚSQUEDA DE PROFESORES EN EL EXCEL
            # ==============================================================
            if query:
                for sheet_idx, nombre_hoja in enumerate(wb.sheetnames):
                    hoja = wb[nombre_hoja]
                    nombre_hoja_lower = nombre_hoja.lower()

                    if 'kawak' in nombre_hoja_lower or 'control' in nombre_hoja_lower:
                        continue

                    filas_muestra = list(hoja.iter_rows(max_row=5, values_only=True))
                    if not filas_muestra:
                        continue

                    texto_cabecera = " ".join([str(c).upper() for f in filas_muestra for c in f if c])

                    # 1. PESTAÑAS DE ZONAS (Ej: 5º - 8º)
                    if '5º - 8º' in nombre_hoja or '9º - 11º' in nombre_hoja or 'ACOMPAÑAMIENTOS' in texto_cabecera:
                        bloque_actual = "Horario por definir"

                        for fila in hoja.iter_rows(min_row=1, values_only=True):
                            if not fila or len(fila) < 8:
                                continue

                            texto_fila_completa = " ".join([str(celda).upper() for celda in fila if celda])

                            if "PRIMER DESCANSO" in texto_fila_completa:
                                bloque_actual = "PRIMER DESCANSO 10:30-10:50"
                            elif "ALMUERZO" in texto_fila_completa:
                                for celda in fila:
                                    if celda and "ALMUERZO" in str(celda).upper():
                                        bloque_actual = str(celda).strip()
                                        break
                            elif "ACOMPAÑAMIENTO PARQUEADERO" in texto_fila_completa or "PARQUEADERO" in texto_fila_completa:
                                bloque_actual = "ACOMPAÑAMIENTO PARQUEADERO"
                            elif "MAÑANA" in texto_fila_completa:
                                bloque_actual = "MAÑANA 7:00-7:15"
                            elif "TARDE" in texto_fila_completa:
                                bloque_actual = "TARDE 3:30-3:45"

                            zona = str(fila[0]).strip() if fila[0] else None
                            if not zona or zona == 'None':
                                zona = str(fila[1]).strip() if fila[1] else None

                            if (not zona or zona.lower() in ['zonas', 'none', 'dia 1', 'días']) and not (
                                    "MAÑANA" in texto_fila_completa or "TARDE" in texto_fila_completa):
                                continue

                            zona_final = zona if zona and zona.lower() != 'none' else "Parqueadero / Túnel"

                            for i in range(2, 8):
                                if i >= len(fila): break
                                profesor_celda = str(fila[i]).strip()

                                if profesor_celda and profesor_celda.lower() != 'none' and coinciden_nombres(query, normalizar_texto(profesor_celda)):
                                    turnos_zonas.append({
                                        'profesor': profesor_celda,
                                        'zona': zona_final,
                                        'dia': f"Día {i - 1}",
                                        'bloque': bloque_actual,
                                        'color_tarjeta': colores[len(turnos_zonas) % len(colores)],
                                        'icono': iconos[0]
                                    })

                    # 2. PESTAÑAS AFL
                    elif es_hoja_afl(nombre_hoja, filas_muestra, texto_cabecera, sheet_idx, rango_afl):

                        todas_las_filas = list(hoja.iter_rows(values_only=True))
                        if len(todas_las_filas) < 3:
                            continue

                        inicio_secciones = []
                        for idx, fila in enumerate(todas_las_filas):
                            if fila_es_ciclo(fila):
                                inicio_secciones.append(idx)

                        if not inicio_secciones:
                            for idx, fila in enumerate(todas_las_filas):
                                if fila_es_grupo(fila) and idx > 0:
                                    anterior = todas_las_filas[idx - 1]
                                    if fila_es_dia(anterior) or fila_es_ciclo(anterior) or any(contiene_mes(celda) for celda in anterior if celda):
                                        inicio_secciones.append(idx - 1)
                                        break

                                if not inicio_secciones:
                                    inicio_secciones = [0]

                        for sec_idx, inicio_seccion in enumerate(inicio_secciones):
                            siguiente_inicio = inicio_secciones[sec_idx + 1] if sec_idx + 1 < len(inicio_secciones) else len(todas_las_filas)

                            fila_cabecera_idx = -1
                            for idx in range(inicio_seccion + 1, siguiente_inicio):
                                fila = todas_las_filas[idx]
                                if not fila:
                                    continue
                                if fila_es_grupo(fila):
                                    fila_cabecera_idx = idx
                                    break

                            if fila_cabecera_idx == -1:
                                continue

                            fila_cabecera = todas_las_filas[fila_cabecera_idx]
                            fila_ciclo = todas_las_filas[inicio_seccion]

                            columnas_grupo = [c_idx for c_idx, celda in enumerate(fila_cabecera)
                                              if celda and 'grupo' in normalizar_texto(celda)]
                            if not columnas_grupo:
                                continue

                            for num_fila in range(fila_cabecera_idx + 1, siguiente_inicio):
                                fila = todas_las_filas[num_fila]
                                if not fila:
                                    continue

                                for i, celda_profesor in enumerate(fila):
                                    if not celda_profesor:
                                        continue
                                    profesor_str = str(celda_profesor).strip()
                                    if not profesor_str or profesor_str.lower() in ['none', ''] or i in columnas_grupo:
                                        continue

                                    if coinciden_nombres(query, normalizar_texto(profesor_str)):
                                        grupo_col_idx = columnas_grupo[0]
                                        for cg in reversed(columnas_grupo):
                                            if i > cg:
                                                grupo_col_idx = cg
                                                break

                                        grupo = None
                                        if grupo_col_idx < len(fila):
                                            grupo = fila[grupo_col_idx]
                                        if not grupo or str(grupo).strip().lower() in ['', 'none']:
                                            for r_idx in range(num_fila, fila_cabecera_idx, -1):
                                                if grupo_col_idx < len(todas_las_filas[r_idx]):
                                                    val = todas_las_filas[r_idx][grupo_col_idx]
                                                    if val and str(val).strip().lower() not in ['', 'none']:
                                                        grupo = val
                                                        break
                                        grupo = str(grupo).strip() if grupo else 'Grupo por definir'

                                        dia_str = 'Día por definir'
                                        if i < len(fila_cabecera) and fila_cabecera[i]:
                                            dia_txt = str(fila_cabecera[i]).strip()
                                            if dia_txt.lower() != 'none':
                                                dia_str = dia_txt

                                        ciclo_str = f'Periodo: {nombre_hoja.strip()}'
                                        if fila_ciclo:
                                            ciclo_val = None
                                            max_col = min(i, len(fila_ciclo) - 1)
                                            for col_busqueda in range(max_col, -1, -1):
                                                v = fila_ciclo[col_busqueda]
                                                if v and str(v).strip().lower() != 'none':
                                                    ciclo_val = str(v).strip()
                                                    break
                                            if ciclo_val:
                                                ciclo_str = ciclo_val

                                        turnos_afl.append({
                                            'profesor': profesor_str,
                                            'zona': f'AFL - {grupo}',
                                            'dia': dia_str,
                                            'bloque': ciclo_str,
                                            'color_tarjeta': colores[len(turnos_afl) % len(colores)],
                                            'icono': iconos[1]
                                        })

                    # 3. PESTAÑAS DE CANCHAS (Para los resultados de búsqueda)
                    elif 'cancha' in nombre_hoja_lower or 'distribución' in texto_cabecera:
                        for fila in hoja.iter_rows(min_row=1, values_only=True):
                            for i in range(len(fila)):
                                celda = str(fila[i]).strip()
                                if celda != 'none' and coinciden_nombres(query, normalizar_texto(celda)):
                                    semana = str(fila[1]).strip() if len(fila) > 1 and fila[1] else "Semana no definida"
                                    if semana.lower() == 'semana': continue

                                    turnos_zonas.append({
                                        'profesor': celda,
                                        'zona': nombre_hoja,
                                        'dia': semana,
                                        'bloque': "Distribución Cancha",
                                        'color_tarjeta': colores[len(turnos_zonas) % len(colores)],
                                        'icono': iconos[2]
                                    })

        except Exception as e:
            print(f"Error procesando el Excel: {e}")

    if turnos_afl:
        turnos_afl[0]['proximo'] = True

    context = {
        'turnos_zonas': turnos_zonas,
        'turnos_afl': turnos_afl,
        'turnos': turnos_zonas + turnos_afl,
        'cancha_data': cancha_data if 'cancha_data' in locals() else [],
    }

    return render(request, 'acompaapp/consulta.html', context)